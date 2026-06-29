#!/usr/bin/env python3
"""Generate a monthly pontaj (timesheet) tab from a template worksheet."""

from __future__ import annotations

import calendar
import math
import random
import sys
from datetime import date
from pathlib import Path

import holidays
import openpyxl
import typer
from openpyxl.worksheet.worksheet import Worksheet

# --- editable defaults ---
DEFAULT_FILE = "Anexa 1_Fisa individuala de pontaj-din 2026 - Toren Dmitri.xlsx"
DEFAULT_TEMPLATE = "May"
TARGET_HOURS: dict[str, int] = {
    "A.2.1.8.1": 48,
    "A.2.1.7.1": 48,
    "A.2.1.10.1": 20,
    "A.2.1.11.1": 52,
}
ALLOWED_DAILY_HOURS: tuple[int, ...] = (4, 6, 8, 10)
DAILY_ROW_START = 19
DAILY_ROW_END = 49

ROMANIAN_MONTHS: dict[int, str] = {
    1: "ianuarie",
    2: "februarie",
    3: "martie",
    4: "aprilie",
    5: "mai",
    6: "iunie",
    7: "iulie",
    8: "august",
    9: "septembrie",
    10: "octombrie",
    11: "noiembrie",
    12: "decembrie",
}

ENGLISH_MONTHS: dict[int, str] = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

SUMMARY_ROWS: dict[str, int] = {
    "A.2.1.8.1": 19,
    "A.2.1.7.1": 20,
    "A.2.1.10.1": 21,
    "A.2.1.11.1": 22,
}

app = typer.Typer(add_completion=False, no_args_is_help=True)


def day_count_bounds(target_hours: int) -> tuple[int, int]:
    return math.ceil(target_hours / 10), math.floor(target_hours / 4)


def is_working_day(day: date, ro_holidays: holidays.HolidayBase) -> bool:
    return day.weekday() < 5 and day not in ro_holidays


def working_days_in_month(year: int, month: int) -> list[date]:
    ro_holidays = holidays.Romania(years=year)
    _, days_in_month = calendar.monthrange(year, month)
    return [
        date(year, month, day)
        for day in range(1, days_in_month + 1)
        if is_working_day(date(year, month, day), ro_holidays)
    ]


def last_working_day(year: int, month: int) -> date:
    working = working_days_in_month(year, month)
    if not working:
        raise ValueError(f"No working days found for {year}-{month:02d}")
    return working[-1]


def activity_titles_from_template(template_ws: Worksheet) -> dict[str, str]:
    titles: dict[str, str] = {}
    for row in range(DAILY_ROW_START, DAILY_ROW_END + 1):
        code = template_ws[f"B{row}"].value
        title = template_ws[f"C{row}"].value
        if isinstance(code, str) and isinstance(title, str) and code not in titles:
            titles[code] = title
    missing = set(TARGET_HOURS) - set(titles)
    if missing:
        raise ValueError(f"Template sheet is missing activity titles for: {sorted(missing)}")
    return titles


def hours_feasible_for_days(day_count: int, target_hours: int) -> bool:
    return day_count * min(ALLOWED_DAILY_HOURS) <= target_hours <= day_count * max(ALLOWED_DAILY_HOURS)


def assign_day_counts(
    targets: dict[str, int],
    working_day_count: int,
) -> dict[str, int]:
    codes = list(targets.keys())
    bounds = {code: day_count_bounds(hours) for code, hours in targets.items()}

    for code, hours in targets.items():
        lo, hi = bounds[code]
        if hi < lo:
            raise ValueError(
                f"Activity {code}: target {hours}h is incompatible with day bounds "
                f"[{lo}, {hi}] (ceil(T/10) .. floor(T/4))"
            )

    total_target = sum(targets.values())
    counts = {
        code: max(
            bounds[code][0],
            min(bounds[code][1], round(targets[code] / total_target * working_day_count)),
        )
        for code in codes
    }

    def total() -> int:
        return sum(counts.values())

    guard = 0
    while total() != working_day_count and guard < 10_000:
        guard += 1
        if total() < working_day_count:
            candidates = [
                code
                for code in codes
                if counts[code] < bounds[code][1]
            ]
            if not candidates:
                break
            code = max(candidates, key=lambda c: targets[c] - counts[c] * 8)
            counts[code] += 1
        else:
            candidates = [
                code
                for code in codes
                if counts[code] > bounds[code][0]
            ]
            if not candidates:
                break
            code = max(candidates, key=lambda c: counts[c] * 8 - targets[c])
            counts[code] -= 1

    if total() != working_day_count:
        raise ValueError(
            f"Cannot assign {working_day_count} working days across targets {targets} "
            f"within per-activity day bounds {bounds}"
        )

    for code in codes:
        if not hours_feasible_for_days(counts[code], targets[code]):
            raise ValueError(
                f"Activity {code}: {counts[code]} days cannot reach {targets[code]}h "
                f"with daily blocks {ALLOWED_DAILY_HOURS}"
            )

    return counts


def adjust_hours_to_target(
    day_hours: list[int],
    target_hours: int,
    rng: random.Random,
) -> None:
    if not day_hours:
        raise ValueError("Cannot allocate hours without assigned days")

    diff = target_hours - sum(day_hours)
    if diff % 2 != 0:
        raise ValueError(
            f"Target {target_hours}h is not reachable with ±2h steps from base 8h days"
        )

    guard = 0
    while diff != 0 and guard < 10_000:
        guard += 1
        progress = False
        order = list(range(len(day_hours)))
        rng.shuffle(order)
        for index in order:
            if diff > 0 and day_hours[index] + 2 <= max(ALLOWED_DAILY_HOURS):
                day_hours[index] += 2
                diff -= 2
                progress = True
                break
            if diff < 0 and day_hours[index] - 2 >= min(ALLOWED_DAILY_HOURS):
                day_hours[index] -= 2
                diff += 2
                progress = True
                break
        if not progress:
            raise ValueError(
                f"Could not adjust daily hours to reach {target_hours}h; "
                f"current blocks {day_hours}"
            )

    if any(hour not in ALLOWED_DAILY_HOURS for hour in day_hours):
        raise ValueError(f"Invalid daily hour blocks produced: {day_hours}")


def has_three_in_a_row(sequence: list[str]) -> bool:
    return any(
        sequence[index] == sequence[index + 1] == sequence[index + 2]
        for index in range(len(sequence) - 2)
    )


def build_activity_sequence(
    day_counts: dict[str, int],
    targets: dict[str, int],
    rng: random.Random,
) -> list[tuple[str, int]]:
    hours_by_code: dict[str, list[int]] = {}
    for code, day_count in day_counts.items():
        blocks = [8] * day_count
        adjust_hours_to_target(blocks, targets[code], rng)
        hours_by_code[code] = blocks

    pairs: list[tuple[str, int]] = []
    for code, blocks in hours_by_code.items():
        pairs.extend((code, hour) for hour in blocks)

    codes_only = [code for code, _hour in pairs]
    for _attempt in range(5_000):
        rng.shuffle(pairs)
        shuffled_codes = [code for code, _hour in pairs]
        if not has_three_in_a_row(shuffled_codes):
            return pairs

    raise ValueError(
        "Could not randomize activity order without 3 identical activities in a row; "
        f"try a different --seed (last sequence: {codes_only})"
    )


def allocate_schedule(
    year: int,
    month: int,
    targets: dict[str, int],
    leave_days: int,
    seed: int,
) -> tuple[list[date], list[tuple[str, int]]]:
    working = working_days_in_month(year, month)
    if leave_days:
        if leave_days >= len(working):
            raise ValueError(
                f"--leave {leave_days} removes all {len(working)} working days in "
                f"{ENGLISH_MONTHS[month]} {year}"
            )
        working = working[:-leave_days]

    working_day_count = len(working)
    total_target = sum(targets.values())
    min_total = working_day_count * min(ALLOWED_DAILY_HOURS)
    max_total = working_day_count * max(ALLOWED_DAILY_HOURS)
    if total_target < min_total or total_target > max_total:
        raise ValueError(
            f"{total_target}h is infeasible for {working_day_count} working days "
            f"(allowed range {min_total}h..{max_total}h with blocks {ALLOWED_DAILY_HOURS})"
        )

    day_counts = assign_day_counts(targets, working_day_count)
    rng = random.Random(seed)
    schedule = build_activity_sequence(day_counts, targets, rng)

    if len(schedule) != working_day_count:
        raise ValueError(
            f"Internal error: schedule length {len(schedule)} != working days {working_day_count}"
        )

    actual_by_code: dict[str, int] = {code: 0 for code in targets}
    for code, hours in schedule:
        actual_by_code[code] += hours
    if actual_by_code != targets:
        raise ValueError(
            f"Internal error: allocated {actual_by_code}, expected {targets}"
        )

    return working, schedule


def clear_daily_rows(ws: Worksheet) -> None:
    for day in range(1, 32):
        row = 18 + day
        ws[f"A{row}"] = float(day)
        for column in "BCDEFGH":
            ws[f"{column}{row}"] = None


def fill_daily_rows(
    ws: Worksheet,
    year: int,
    month: int,
    titles: dict[str, str],
    schedule_by_date: dict[date, tuple[str, int]],
) -> None:
    ro_holidays = holidays.Romania(years=year)
    _, days_in_month = calendar.monthrange(year, month)

    for day in range(1, 32):
        row = 18 + day
        ws[f"A{row}"] = float(day)
        ws[f"B{row}"] = None
        ws[f"C{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None
        ws[f"F{row}"] = None
        ws[f"G{row}"] = None
        ws[f"H{row}"] = None

        if day > days_in_month:
            continue

        current = date(year, month, day)
        if not is_working_day(current, ro_holidays):
            continue

        if current not in schedule_by_date:
            continue

        code, hours = schedule_by_date[current]
        ws[f"B{row}"] = code
        ws[f"C{row}"] = titles[code]
        ws[f"E{row}"] = float(hours)
        ws[f"H{row}"] = f"=SUM(E{row}:G{row})"


def update_summary_formulas(ws: Worksheet) -> None:
    for code, row in SUMMARY_ROWS.items():
        ws[f"L{row}"] = f'=SUMIF($B${DAILY_ROW_START}:$B${DAILY_ROW_END},"{code}",$E${DAILY_ROW_START}:$E${DAILY_ROW_END})'


def update_totals_and_signatures(ws: Worksheet, year: int, month: int) -> None:
    ws["E50"] = f"=SUM(E{DAILY_ROW_START}:E{DAILY_ROW_END})"
    ws["H50"] = f"=SUM(H{DAILY_ROW_START}:H{DAILY_ROW_END})"
    last_day = last_working_day(year, month).strftime("%d.%m.%Y")
    ws["B55"] = last_day
    ws["F55"] = last_day


def clone_template_sheet(
    workbook: openpyxl.Workbook,
    template_name: str,
    target_name: str,
) -> Worksheet:
    if template_name not in workbook.sheetnames:
        raise ValueError(f"Template sheet {template_name!r} not found in workbook")
    if target_name in workbook.sheetnames:
        del workbook[target_name]
    copied = workbook.copy_worksheet(workbook[template_name])
    copied.title = target_name
    return copied


def build_pontaj_sheet(
    workbook: openpyxl.Workbook,
    template_name: str,
    year: int,
    month: int,
    leave_days: int,
    seed: int,
) -> Worksheet:
    template_ws = workbook[template_name]
    titles = activity_titles_from_template(template_ws)
    target_name = ENGLISH_MONTHS[month]
    ws = clone_template_sheet(workbook, template_name, target_name)

    ws["A8"] = f"{ROMANIAN_MONTHS[month]} {year}"
    clear_daily_rows(ws)

    working, schedule = allocate_schedule(year, month, TARGET_HOURS, leave_days, seed)
    schedule_by_date = dict(zip(working, schedule, strict=True))
    fill_daily_rows(ws, year, month, titles, schedule_by_date)
    update_summary_formulas(ws)
    update_totals_and_signatures(ws, year, month)

    return ws


def print_summary(
    year: int,
    month: int,
    leave_days: int,
    seed: int,
    sheet_name: str,
    output_path: Path,
) -> None:
    working = working_days_in_month(year, month)
    if leave_days:
        working = working[:-leave_days]

    _working, schedule = allocate_schedule(year, month, TARGET_HOURS, leave_days, seed)
    actual_by_code: dict[str, int] = {code: 0 for code in TARGET_HOURS}
    for code, hours in schedule:
        actual_by_code[code] += hours

    print(f"Created sheet {sheet_name!r} in {output_path}")
    print(f"Working days: {len(working)} (leave days excluded: {leave_days})")
    print(f"Total hours: {sum(actual_by_code.values())} (target {sum(TARGET_HOURS.values())})")
    print("Target vs actual by activity:")
    for code in TARGET_HOURS:
        print(f"  {code}: target {TARGET_HOURS[code]}h, actual {actual_by_code[code]}h")
    print("Note: Excel recalculates formulas when the workbook is opened.")


@app.command()
def main(
    year: int = typer.Argument(..., help="Target year, e.g. 2026"),
    month: int = typer.Argument(..., min=1, max=12, help="Target month (1-12)"),
    seed: int | None = typer.Option(
        None,
        "--seed",
        help="Random seed (default: year*100+month)",
    ),
    leave: int = typer.Option(
        0,
        "--leave",
        min=0,
        help="Number of trailing working days to leave unallocated",
    ),
    file: Path = typer.Option(
        Path(DEFAULT_FILE),
        "--file",
        help="Input workbook path",
    ),
    template: str = typer.Option(
        DEFAULT_TEMPLATE,
        "--template",
        help="Template worksheet name to clone",
    ),
    out: Path | None = typer.Option(
        None,
        "--out",
        help="Output workbook path (default: overwrite --file)",
    ),
) -> None:
    if not file.exists():
        raise typer.BadParameter(f"Workbook not found: {file}")

    effective_seed = seed if seed is not None else year * 100 + month
    output_path = out if out is not None else file
    sheet_name = ENGLISH_MONTHS[month]

    workbook = openpyxl.load_workbook(file)
    build_pontaj_sheet(workbook, template, year, month, leave, effective_seed)
    workbook.save(output_path)

    print_summary(year, month, leave, effective_seed, sheet_name, output_path)


if __name__ == "__main__":
    try:
        app()
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
