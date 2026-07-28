#!/usr/bin/env python3
"""Batch functional annotation for the July prioritized variant set.

Reads candidate variants, queries GTEx v8 eQTLs (brain + whole blood) and
Ensembl VEP (consequences, SIFT/PolyPhen, HGVS), joins pre-computed
AlphaGenome / AlphaMissense score matrices, and writes a three-sheet Excel
workbook under ``outputs/``.

Typical usage::

    uv run python run_july_annotation_pipeline.py
    uv run python run_july_annotation_pipeline.py --cache-only -v

See also:
    docs/JULY_ANNOTATION_PIPELINE.md
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
import time
from pathlib import Path
from typing import Any, Literal
from urllib.parse import quote

import polars as pl
import requests
import typer
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

ScoreKind = Literal["alphagenome", "alphamissense"]
JsonDict = dict[str, Any]
JsonPayload = Any

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
INPUT_MANIFEST = REPO_ROOT / "INPUT_MANIFEST.md"

DEFAULT_VARIANTS = REPO_ROOT / "data" / "processed" / "variants_47_input.csv"
DEFAULT_ALPHAGENOME = REPO_ROOT / "data" / "scores" / "alphagenome_raw.parquet"
DEFAULT_ALPHAMISSENSE = REPO_ROOT / "data" / "scores" / "alphamissense_raw.parquet"
DEFAULT_OUTPUT = REPO_ROOT / "outputs" / "Supplementary_Table_1_Annotated_Variants.xlsx"
DEFAULT_CACHE_DIR = REPO_ROOT / "data" / "cache" / "july_annotation"
DEFAULT_LOCAL_VEP = REPO_ROOT / "data" / "processed" / "vep_local.jsonl"

REQUIRED_VARIANT_COLS = ("chrom", "pos", "ref", "alt", "rsid", "gene_symbol")
EXPECTED_VARIANT_COUNT = 47
JULY_MANIFEST_REQUIRED: tuple[str, ...] = (
    "data/processed/variants_47_input.csv",
    "data/scores/alphagenome_raw.parquet",
    "data/scores/alphamissense_raw.parquet",
)

GTEX_API_BASE = "https://gtexportal.org/api/v2"
GTEX_DATASET_ID = "gtex_v8"
ENSEMBL_REST = "https://rest.ensembl.org"

TARGET_TISSUES: list[str] = [
    "Brain_Amygdala",
    "Brain_Anterior_cingulate_cortex_BA24",
    "Brain_Caudate_basal_ganglia",
    "Brain_Cerebellar_Hemisphere",
    "Brain_Cerebellum",
    "Brain_Cortex",
    "Brain_Frontal_Cortex_BA9",
    "Brain_Hippocampus",
    "Brain_Hypothalamus",
    "Brain_Nucleus_accumbens_basal_ganglia",
    "Brain_Putamen_basal_ganglia",
    "Brain_Spinal_cord_cervical_c-1",
    "Brain_Substantia_nigra",
    "Whole_Blood",
]

IMPACT_RANK: dict[str, int] = {
    "HIGH": 0,
    "MODERATE": 1,
    "LOW": 2,
    "MODIFIER": 3,
}

HIGH_IMPACT_LEVELS = frozenset({"HIGH", "MODERATE"})
ALPHAMISSENSE_HIGH_THRESHOLD = 0.5

REQUEST_TIMEOUT_SEC = 30.0
GTEX_DELAY_SEC = 0.5
VEP_DELAY_SEC = 0.34
MAX_RETRIES = 4
ITEMS_PER_PAGE = 250

RSID_RE = re.compile(r"^rs\d+$", re.IGNORECASE)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")

app = typer.Typer(add_completion=False, help=__doc__, no_args_is_help=False)


def verify_july_input_manifest(
    manifest_path: Path = INPUT_MANIFEST,
    *,
    repo_root: Path = REPO_ROOT,
) -> None:
    """Confirm ``INPUT_MANIFEST.md`` exists and July required inputs are on disk.

    Args:
        manifest_path: Path to the repository input manifest.
        repo_root: Repository root used to resolve relative required paths.

    Raises:
        FileNotFoundError: If the manifest or any required July input is missing.
    """
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"INPUT_MANIFEST.md not found at {manifest_path}. "
            "Read/create the Activity A.2.1.8.1 July annotation section before production runs."
        )
    text = manifest_path.read_text(encoding="utf-8")
    if "variants_47_input.csv" not in text:
        raise FileNotFoundError(
            f"{manifest_path} was read but does not document variants_47_input.csv "
            "(Activity A.2.1.8.1 July annotation inputs)."
        )
    missing = [
        rel for rel in JULY_MANIFEST_REQUIRED if not (repo_root / rel).is_file()
    ]
    if missing:
        raise FileNotFoundError(
            "Required July annotation input(s) listed in INPUT_MANIFEST.md are missing:\n  - "
            + "\n  - ".join(missing)
        )
    logger.info("INPUT_MANIFEST.md verified for July annotation ({} required files)", len(JULY_MANIFEST_REQUIRED))


def configure_logging(verbose: bool) -> None:
    """Configure loguru sinks for console output.

    Args:
        verbose: When True, emit DEBUG-level messages; otherwise INFO.
    """
    logger.remove()
    level = "DEBUG" if verbose else "INFO"
    logger.add(
        sys.stderr,
        level=level,
        format=(
            "<green>{time:YYYY-MM-DD HH:mm:ss}</green> | "
            "<level>{level: <8}</level> | "
            "<cyan>{message}</cyan>"
        ),
    )


def normalize_chrom(value: object) -> str:
    """Return chromosome without a ``chr`` prefix.

    Args:
        value: Raw chromosome token (e.g. ``chr2``, ``2``).

    Returns:
        Chromosome string without a leading ``chr`` prefix.
    """
    text = str(value).strip()
    if text.lower().startswith("chr"):
        return text[3:]
    return text


def gtex_chromosome(value: object) -> str:
    """Format chromosome as GTEx expects (``chrN``).

    Args:
        value: Raw chromosome token.

    Returns:
        GTEx-style chromosome identifier such as ``chr2``.
    """
    chrom = normalize_chrom(value)
    return f"chr{chrom}"


def normalize_rsid(raw: object) -> str | None:
    """Return a normalized rsID, or None when the value is not rs-formatted.

    Args:
        raw: Candidate rsID value from a table cell.

    Returns:
        Trimmed rsID string, or None if missing / malformed.
    """
    if raw is None:
        return None
    token = str(raw).strip()
    if not token or token.lower() in {"nan", "none", "null"}:
        return None
    if not RSID_RE.match(token):
        return None
    return token


def variant_key(chrom: object, pos: object, ref: object, alt: object) -> str:
    """Build a locus key used for joining score matrices.

    Multi-allelic ``alt`` values are reduced to the first allele so joins stay
    1:1 with the prioritized-variant table.

    Args:
        chrom: Chromosome.
        pos: 1-based GRCh38 position.
        ref: Reference allele.
        alt: Alternate allele (comma-separated allowed).

    Returns:
        Canonical key ``{chrom}:{pos}:{REF}:{ALT}``.
    """
    primary_alt = str(alt).split(",")[0].strip()
    return f"{normalize_chrom(chrom)}:{int(pos)}:{str(ref).strip().upper()}:{primary_alt.upper()}"


def cache_path(cache_dir: Path, kind: str, key: str) -> Path:
    """Return a filesystem-safe JSON cache path.

    Long keys are truncated and hashed so paths stay within OS limits.

    Args:
        cache_dir: Root cache directory.
        kind: Cache namespace prefix (e.g. ``vep_id``, ``gtex_eqtl``).
        key: Query-specific identifier.

    Returns:
        Path to the JSON cache file for this query.
    """
    safe = quote(key, safe="")
    if len(safe) > 180:
        # Truncate + content hash avoids ENAMETOOLONG while keeping uniqueness.
        digest = hashlib.sha256(key.encode("utf-8")).hexdigest()[:24]
        safe = f"{safe[:80]}__{digest}"
    return cache_dir / f"{kind}_{safe}.json"


def pace(last_request_end: float | None, min_interval_sec: float) -> float:
    """Sleep if needed to respect API pacing; return a fresh monotonic timestamp.

    Args:
        last_request_end: Monotonic time when the previous request finished.
        min_interval_sec: Minimum seconds between live HTTP calls.

    Returns:
        Current ``time.monotonic()`` after any required sleep.
    """
    now = time.monotonic()
    if last_request_end is not None:
        wait = min_interval_sec - (now - last_request_end)
        if wait > 0:
            time.sleep(wait)
    return time.monotonic()


def fetch_json_with_retry(
    session: requests.Session,
    url: str,
    *,
    params: list[tuple[str, str]] | dict[str, str] | None,
    cache_file: Path,
    min_interval_sec: float,
    max_retries: int,
    last_request_end: float | None,
    cache_only: bool,
    allow_404: bool = False,
) -> tuple[JsonPayload | None, float | None, bool]:
    """GET JSON with disk cache, exponential backoff on 429/503, and pacing.

    Args:
        session: Shared ``requests.Session``.
        url: Absolute request URL.
        params: Query string parameters.
        cache_file: On-disk JSON cache path for this query.
        min_interval_sec: Polite delay between live calls.
        max_retries: Maximum retries for rate-limit / transient errors.
        last_request_end: Monotonic timestamp of the previous live call.
        cache_only: If True, never hit the network on cache miss.
        allow_404: If True, treat HTTP 404 as a soft miss (return None).

    Returns:
        Tuple of ``(payload, last_request_end, from_cache)``.
    """
    if cache_file.is_file():
        with cache_file.open(encoding="utf-8") as handle:
            return json.load(handle), last_request_end, True

    if cache_only:
        logger.warning("cache miss (cache-only mode): {}", cache_file.name)
        return None, last_request_end, False

    attempt = 0
    while True:
        attempt += 1
        last_request_end = pace(last_request_end, min_interval_sec)
        response: requests.Response | None = None
        try:
            response = session.get(url, params=params, timeout=REQUEST_TIMEOUT_SEC)
        except requests.RequestException as exc:
            last_request_end = time.monotonic()
            logger.warning("request error for {}: {}", url, exc)
            if attempt > max_retries:
                return None, last_request_end, False
            time.sleep(min_interval_sec * (2 ** (attempt - 1)))
            continue
        last_request_end = time.monotonic()

        if response.status_code == 404 and allow_404:
            return None, last_request_end, False

        if response.status_code in {429, 503}:
            retry_after = response.headers.get("Retry-After")
            try:
                sleep_s = float(retry_after) if retry_after is not None else min_interval_sec * (
                    2 ** (attempt - 1)
                )
            except ValueError:
                sleep_s = min_interval_sec * (2 ** (attempt - 1))
            sleep_s = max(sleep_s, min_interval_sec)
            if attempt > max_retries:
                logger.error("giving up on {} after HTTP {}", url, response.status_code)
                return None, last_request_end, False
            logger.warning("HTTP {}; retry in {:.2f}s ({})", response.status_code, sleep_s, url)
            time.sleep(sleep_s)
            continue

        if not response.ok:
            logger.error("HTTP {} for {}: {}", response.status_code, url, response.text[:300])
            return None, last_request_end, False

        payload: JsonPayload = response.json()
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        with cache_file.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2)
        return payload, last_request_end, False


def load_prioritized_variants(path: Path) -> pl.DataFrame:
    """Load and validate the prioritized variant CSV.

    Normalizes alleles to uppercase, strips ``chr`` prefixes, keeps the first
    alternate allele for multi-allelic rows, and adds ``variant_key``.

    Args:
        path: CSV path with required columns in ``REQUIRED_VARIANT_COLS``.

    Returns:
        Polars DataFrame ready for annotation joins.

    Raises:
        FileNotFoundError: If ``path`` does not exist.
        ValueError: If required columns are missing.
    """
    if not path.is_file():
        raise FileNotFoundError(f"Prioritized variant list not found: {path}")

    frame = pl.read_csv(path, infer_schema_length=10_000)
    missing = [col for col in REQUIRED_VARIANT_COLS if col not in frame.columns]
    if missing:
        raise ValueError(
            f"Missing required columns in {path.name}: {missing}; available: {frame.columns}"
        )

    frame = frame.with_columns(
        pl.col("chrom")
        .cast(pl.Utf8)
        .str.replace(r"(?i)^chr", "")
        .alias("chrom"),
        pl.col("pos").cast(pl.Int64),
        pl.col("ref").cast(pl.Utf8).str.to_uppercase(),
        # Keep a single alt allele so GTEx / score joins remain unambiguous.
        pl.col("alt")
        .cast(pl.Utf8)
        .str.split(",")
        .list.first()
        .str.to_uppercase()
        .alias("alt"),
        pl.col("rsid").cast(pl.Utf8),
        pl.col("gene_symbol").cast(pl.Utf8),
    )
    frame = frame.with_columns(
        (
            pl.col("chrom")
            + pl.lit(":")
            + pl.col("pos").cast(pl.Utf8)
            + pl.lit(":")
            + pl.col("ref")
            + pl.lit(":")
            + pl.col("alt")
        ).alias("variant_key")
    )
    logger.info("Loaded {} prioritized variants from {}", frame.height, path)
    if path.resolve() == DEFAULT_VARIANTS.resolve() or path.name == "variants_47_input.csv":
        if frame.height != EXPECTED_VARIANT_COUNT:
            raise ValueError(
                f"Expected {EXPECTED_VARIANT_COUNT} production variants in {path.name}, "
                f"found {frame.height}. Refusing truncated/demo input."
            )
    return frame


def _rename_score_columns(frame: pl.DataFrame, mapping: dict[str, str]) -> pl.DataFrame:
    """Rename known aliases onto canonical score column names when present.

    Args:
        frame: Score matrix DataFrame.
        mapping: Source-column → canonical-column rename map.

    Returns:
        Frame with applicable renames applied (no-op when none match).
    """
    renames = {
        src: dst for src, dst in mapping.items() if src in frame.columns and dst not in frame.columns
    }
    return frame.rename(renames) if renames else frame


def load_score_matrix(path: Path, kind: ScoreKind) -> pl.DataFrame:
    """Load AlphaGenome or AlphaMissense parquet and normalize join keys.

    Accepts either locus columns (``chrom``, ``pos``, ``ref``, ``alt``) or
    ``rsid``. Alias columns from earlier analysis exports are remapped to the
    canonical ``alphagenome_*`` / ``alphamissense_*`` names.

    Args:
        path: Parquet score matrix path.
        kind: Which score family to load.

    Returns:
        Normalized score DataFrame with ``variant_key`` and/or ``rsid``.

    Raises:
        FileNotFoundError: If ``path`` does not exist.
        ValueError: If neither locus nor rsID columns are present.
    """
    if not path.is_file():
        raise FileNotFoundError(f"{kind} score matrix not found: {path}")

    frame = pl.read_parquet(path)
    if kind == "alphagenome":
        frame = _rename_score_columns(
            frame,
            {
                "ref_score": "alphagenome_ref_score",
                "alt_score": "alphagenome_alt_score",
                "diff": "alphagenome_diff",
                "perc_change": "alphagenome_perc_change",
                "abs_perc_change": "alphagenome_abs_perc_change",
                "snp": "rsid",
                "gene": "gene_symbol",
            },
        )
    else:
        frame = _rename_score_columns(
            frame,
            {
                "am_score": "alphamissense_score",
                "alphamissense": "alphamissense_score",
                "am_class": "alphamissense_class",
                "alphamissense_pred": "alphamissense_class",
                "snp": "rsid",
            },
        )

    if "rsid" in frame.columns:
        frame = frame.with_columns(pl.col("rsid").cast(pl.Utf8))

    if all(col in frame.columns for col in ("chrom", "pos", "ref", "alt")):
        frame = frame.with_columns(
            pl.col("chrom").cast(pl.Utf8).str.replace(r"(?i)^chr", "").alias("chrom"),
            pl.col("pos").cast(pl.Int64),
            pl.col("ref").cast(pl.Utf8).str.to_uppercase(),
            pl.col("alt")
            .cast(pl.Utf8)
            .str.split(",")
            .list.first()
            .str.to_uppercase()
            .alias("alt"),
        )
        frame = frame.with_columns(
            (
                pl.col("chrom")
                + pl.lit(":")
                + pl.col("pos").cast(pl.Utf8)
                + pl.lit(":")
                + pl.col("ref")
                + pl.lit(":")
                + pl.col("alt")
            ).alias("variant_key")
        )
    elif "rsid" not in frame.columns:
        raise ValueError(
            f"{kind} matrix at {path} needs either (chrom,pos,ref,alt) or rsid; "
            f"columns={frame.columns}"
        )

    logger.info("Loaded {} {} score rows from {}", frame.height, kind, path)
    return frame


def join_scores(
    variants: pl.DataFrame,
    alphagenome: pl.DataFrame,
    alphamissense: pl.DataFrame,
) -> pl.DataFrame:
    """Left-join AlphaGenome and AlphaMissense scores onto the variant table.

    Prefers locus ``variant_key`` joins; falls back to ``rsid`` when score
    matrices lack coordinates.

    Args:
        variants: Prioritized variant table.
        alphagenome: Normalized AlphaGenome scores.
        alphamissense: Normalized AlphaMissense scores.

    Returns:
        Variant rows with attached score columns (nulls allowed).
    """

    def _score_subset(scores: pl.DataFrame, keep_prefixes: tuple[str, ...]) -> pl.DataFrame:
        """Keep join keys plus score columns matching ``keep_prefixes``."""
        keep = [col for col in scores.columns if col.startswith(keep_prefixes) or col in {
            "variant_key",
            "rsid",
        }]
        # Prefer locus join when available; otherwise rsid.
        if "variant_key" in scores.columns:
            cols = ["variant_key", *[c for c in keep if c not in {"variant_key", "rsid"}]]
            return scores.select([c for c in cols if c in scores.columns]).unique(
                subset=["variant_key"], keep="first"
            )
        cols = ["rsid", *[c for c in keep if c != "rsid"]]
        return scores.select([c for c in cols if c in scores.columns]).unique(
            subset=["rsid"], keep="first"
        )

    ag = _score_subset(alphagenome, ("alphagenome_",))
    am = _score_subset(alphamissense, ("alphamissense_",))

    if "variant_key" in ag.columns:
        merged = variants.join(ag, on="variant_key", how="left")
    else:
        merged = variants.join(ag, on="rsid", how="left")

    if "variant_key" in am.columns and "variant_key" in merged.columns:
        # Drop score-frame rsid to avoid duplicate column names after join.
        am_cols = [c for c in am.columns if c != "rsid" or "rsid" not in merged.columns]
        merged = merged.join(am.select(am_cols), on="variant_key", how="left")
    else:
        am_cols = [c for c in am.columns if c != "variant_key"]
        merged = merged.join(am.select(am_cols), on="rsid", how="left")

    return merged


def load_local_vep(path: Path) -> dict[str, list[JsonDict]]:
    """Load optional local VEP JSONL keyed by rsID or variant_key.

    Each non-empty line must be a JSON object. The VEP payload may live under
    ``payload``, ``vep``, or be the object itself.

    Args:
        path: JSONL file path. Missing file yields an empty mapping.

    Returns:
        Mapping from lookup keys to VEP payload lists.
    """
    if not path.is_file():
        return {}
    by_key: dict[str, list[JsonDict]] = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            payload = record.get("payload") or record.get("vep") or record
            if isinstance(payload, dict):
                payload_list = [payload]
            elif isinstance(payload, list):
                payload_list = payload
            else:
                continue
            for key_name in ("rsid", "id", "variant_key", "input"):
                key = record.get(key_name)
                if key:
                    by_key[str(key)] = payload_list
            if payload_list and payload_list[0].get("id"):
                by_key[str(payload_list[0]["id"])] = payload_list
    logger.info("Loaded {} local VEP records from {}", len(by_key), path)
    return by_key


def pick_sift_polyphen(
    transcripts: list[JsonDict],
) -> tuple[str, str, float | None, float | None]:
    """Pick SIFT/PolyPhen labels and scores from the highest-impact transcript.

    Transcripts are ranked by VEP impact (HIGH → MODIFIER). The first transcript
    that carries any SIFT/PolyPhen field wins.

    Args:
        transcripts: VEP ``transcript_consequences`` list.

    Returns:
        Tuple of ``(sift_prediction, polyphen_prediction, sift_score,
        polyphen_score)``. Empty strings / None when unavailable.
    """
    ranked = sorted(
        transcripts,
        key=lambda tc: IMPACT_RANK.get(str(tc.get("impact", "MODIFIER")), 99),
    )
    for tc in ranked:
        sift_pred = tc.get("sift_prediction") or tc.get("sift")
        poly_pred = tc.get("polyphen_prediction") or tc.get("polyphen")
        sift_score = tc.get("sift_score")
        poly_score = tc.get("polyphen_score")
        if sift_pred or poly_pred or sift_score is not None or poly_score is not None:
            return (
                str(sift_pred or ""),
                str(poly_pred or ""),
                float(sift_score) if sift_score is not None else None,
                float(poly_score) if poly_score is not None else None,
            )
    return "", "", None, None


def pick_primary_transcript(transcripts: list[JsonDict]) -> JsonDict | None:
    """Select the canonical / highest-impact transcript consequence.

    Args:
        transcripts: VEP ``transcript_consequences`` list.

    Returns:
        Preferred transcript dict, or None when the list is empty.
    """
    if not transcripts:
        return None
    # Canonical flag first, then severity — matches manuscript “primary” transcript.
    ranked = sorted(
        transcripts,
        key=lambda tc: (
            0 if tc.get("canonical") in {1, True, "1"} else 1,
            IMPACT_RANK.get(str(tc.get("impact", "MODIFIER")), 99),
        ),
    )
    return ranked[0]


def extract_vep_fields(payload: list[JsonDict] | None) -> JsonDict:
    """Map a VEP JSON payload to flat annotation fields.

    Args:
        payload: VEP REST response list (one variant), or None.

    Returns:
        Flat dict with consequence, impact, SIFT/PolyPhen, and HGVS fields.
    """
    empty: JsonDict = {
        "vep_most_severe_consequence": None,
        "vep_impact": None,
        "vep_transcript_id": None,
        "vep_gene_symbol": None,
        "vep_consequence_terms": None,
        "sift_prediction": None,
        "sift_score": None,
        "polyphen_prediction": None,
        "polyphen_score": None,
        "hgvsg": None,
        "hgvsc": None,
        "hgvsp": None,
    }
    if not payload:
        return empty

    variant = payload[0]
    transcripts = variant.get("transcript_consequences") or []
    if not isinstance(transcripts, list):
        transcripts = []

    primary = pick_primary_transcript(transcripts)
    sift_pred, poly_pred, sift_score, poly_score = pick_sift_polyphen(transcripts)

    impact = None
    consequence_terms: list[str] = []
    transcript_id = None
    gene_symbol = None
    hgvsc = None
    hgvsp = None
    if primary is not None:
        impact = primary.get("impact")
        consequence_terms = [str(t) for t in (primary.get("consequence_terms") or [])]
        transcript_id = primary.get("transcript_id")
        gene_symbol = primary.get("gene_symbol")
        hgvsc = primary.get("hgvsc")
        hgvsp = primary.get("hgvsp")

    # Prefer top-level HGVS-g when present; otherwise scan transcripts.
    hgvsg = variant.get("hgvsg")
    if not hgvsg:
        for tc in transcripts:
            if tc.get("hgvsg"):
                hgvsg = tc["hgvsg"]
                break

    if impact is None and transcripts:
        impacts = [str(tc.get("impact", "")) for tc in transcripts if tc.get("impact")]
        impact = min(impacts, key=lambda x: IMPACT_RANK.get(x, 99)) if impacts else None

    return {
        "vep_most_severe_consequence": variant.get("most_severe_consequence"),
        "vep_impact": impact,
        "vep_transcript_id": transcript_id,
        "vep_gene_symbol": gene_symbol,
        "vep_consequence_terms": ";".join(consequence_terms) if consequence_terms else None,
        "sift_prediction": sift_pred or None,
        "sift_score": sift_score,
        "polyphen_prediction": poly_pred or None,
        "polyphen_score": poly_score,
        "hgvsg": hgvsg,
        "hgvsc": hgvsc,
        "hgvsp": hgvsp,
    }


def fetch_vep_for_variant(
    session: requests.Session,
    row: JsonDict,
    cache_dir: Path,
    local_vep: dict[str, list[JsonDict]],
    last_request_end: float | None,
    cache_only: bool,
) -> tuple[JsonDict, float | None]:
    """Query Ensembl VEP by rsID, falling back to region/allele when needed.

    Lookup order: local JSONL → VEP-by-id → VEP-by-region. Region fallback is
    required when an rsID is missing or Ensembl returns 404 for the id endpoint.

    Args:
        session: HTTP session.
        row: Variant dict including chrom/pos/ref/alt/rsid/variant_key.
        cache_dir: API response cache directory.
        local_vep: Optional pre-loaded local VEP payloads.
        last_request_end: Monotonic timestamp of the previous live request.
        cache_only: Disable live Ensembl calls on cache miss.

    Returns:
        Tuple of ``(flat_vep_fields, last_request_end)``.
    """
    rsid = normalize_rsid(row.get("rsid"))
    chrom = normalize_chrom(row["chrom"])
    pos = int(row["pos"])
    alt = str(row["alt"])
    vkey = str(row["variant_key"])

    for key in filter(None, [rsid, vkey, row.get("rsid")]):
        if key in local_vep:
            return extract_vep_fields(local_vep[key]), last_request_end

    payload: list[JsonDict] | None = None
    if rsid:
        url = (
            f"{ENSEMBL_REST}/vep/human/id/{quote(rsid, safe='')}"
            f"?content-type=application/json&hgvs=1&canonical=1"
        )
        cache_file = cache_path(cache_dir, "vep_id", rsid.lower())
        raw, last_request_end, _ = fetch_json_with_retry(
            session,
            url,
            params=None,
            cache_file=cache_file,
            min_interval_sec=VEP_DELAY_SEC,
            max_retries=MAX_RETRIES,
            last_request_end=last_request_end,
            cache_only=cache_only,
            allow_404=True,
        )
        if isinstance(raw, list):
            payload = raw
        elif isinstance(raw, dict):
            payload = [raw]

    if payload is None:
        # Region endpoint encodes ALT in the path: chrom:pos-pos/ALT
        region = f"{chrom}:{pos}-{pos}/{quote(alt, safe='')}"
        url = (
            f"{ENSEMBL_REST}/vep/human/region/{region}"
            f"?content-type=application/json&hgvs=1&canonical=1"
        )
        cache_file = cache_path(cache_dir, "vep_region", f"{chrom}_{pos}_{alt}")
        raw, last_request_end, _ = fetch_json_with_retry(
            session,
            url,
            params=None,
            cache_file=cache_file,
            min_interval_sec=VEP_DELAY_SEC,
            max_retries=MAX_RETRIES,
            last_request_end=last_request_end,
            cache_only=cache_only,
            allow_404=True,
        )
        if isinstance(raw, list):
            payload = raw
        elif isinstance(raw, dict):
            payload = [raw]

    return extract_vep_fields(payload), last_request_end


def parse_gtex_variant(record: JsonDict) -> JsonDict:
    """Extract GTEx variant identity fields.

    Args:
        record: One object from GTEx ``dataset/variant`` ``data``.

    Returns:
        Dict with ``variantId``, ``snpId``, coordinates, and alleles.
    """
    return {
        "variantId": str(record.get("variantId", "")),
        "snpId": str(record.get("snpId", "")),
        "chromosome": str(record.get("chromosome", "")),
        "pos": int(record.get("pos") or 0),
        "ref": str(record.get("ref", "")),
        "alt": str(record.get("alt", "")),
    }


def resolve_gtex_variant(
    session: requests.Session,
    row: JsonDict,
    cache_dir: Path,
    last_request_end: float | None,
    cache_only: bool,
) -> tuple[JsonDict | None, float | None]:
    """Resolve a GTEx variantId via rsID, then chrom/pos fallback.

    Args:
        session: HTTP session.
        row: Variant dict with coordinates / rsID.
        cache_dir: API cache directory.
        last_request_end: Previous live-request monotonic timestamp.
        cache_only: Disable live GTEx calls on cache miss.

    Returns:
        Tuple of ``(gtex_variant_or_None, last_request_end)``.
    """
    rsid = normalize_rsid(row.get("rsid"))
    chrom = gtex_chromosome(row["chrom"])
    pos = str(int(row["pos"]))

    if rsid:
        cache_file = cache_path(cache_dir, "gtex_variant", rsid.lower())
        payload, last_request_end, _ = fetch_json_with_retry(
            session,
            f"{GTEX_API_BASE}/dataset/variant",
            params=[
                ("snpId", rsid),
                ("datasetId", GTEX_DATASET_ID),
                ("itemsPerPage", "250"),
            ],
            cache_file=cache_file,
            min_interval_sec=GTEX_DELAY_SEC,
            max_retries=MAX_RETRIES,
            last_request_end=last_request_end,
            cache_only=cache_only,
        )
        if payload and payload.get("data"):
            return parse_gtex_variant(payload["data"][0]), last_request_end

    loc_cache = cache_path(cache_dir, "gtex_variant_loc", f"{chrom}_{pos}")
    loc_payload, last_request_end, _ = fetch_json_with_retry(
        session,
        f"{GTEX_API_BASE}/dataset/variant",
        params=[
            ("chromosome", chrom),
            ("pos", pos),
            ("datasetId", GTEX_DATASET_ID),
            ("itemsPerPage", "250"),
        ],
        cache_file=loc_cache,
        min_interval_sec=GTEX_DELAY_SEC,
        max_retries=MAX_RETRIES,
        last_request_end=last_request_end,
        cache_only=cache_only,
    )
    if not loc_payload or not loc_payload.get("data"):
        return None, last_request_end

    ref = str(row["ref"]).upper()
    alt = str(row["alt"]).upper()
    for record in loc_payload["data"]:
        variant = parse_gtex_variant(record)
        if rsid and variant["snpId"].lower() == rsid.lower():
            return variant, last_request_end
        if (
            variant["pos"] == int(row["pos"])
            and variant["ref"].upper() == ref
            and variant["alt"].upper() == alt
        ):
            return variant, last_request_end
    return parse_gtex_variant(loc_payload["data"][0]), last_request_end


def fetch_gtex_eqtls(
    session: requests.Session,
    variant_id: str,
    cache_dir: Path,
    last_request_end: float | None,
    cache_only: bool,
) -> tuple[list[JsonDict], float | None]:
    """Fetch significant single-tissue eQTLs for brain regions + whole blood.

    Pages through GTEx ``association/singleTissueEqtl`` for ``TARGET_TISSUES``
    and caches the aggregated hit list.

    Args:
        session: HTTP session.
        variant_id: GTEx ``variantId`` (not an rsID).
        cache_dir: API cache directory.
        last_request_end: Previous live-request monotonic timestamp.
        cache_only: Disable live GTEx calls on cache miss.

    Returns:
        Tuple of ``(eqtl_hit_dicts, last_request_end)``.
    """
    tissue_key = "_".join(TARGET_TISSUES)
    aggregate_cache = cache_path(cache_dir, "gtex_eqtl", f"{variant_id}__{tissue_key}")
    if aggregate_cache.is_file():
        with aggregate_cache.open(encoding="utf-8") as handle:
            cached = json.load(handle)
        return cached.get("hits", []), last_request_end

    hits: list[JsonDict] = []
    page = 0
    number_of_pages = 1
    while page < number_of_pages:
        params: list[tuple[str, str]] = [
            ("variantId", variant_id),
            ("datasetId", GTEX_DATASET_ID),
            ("itemsPerPage", str(ITEMS_PER_PAGE)),
            ("page", str(page)),
        ]
        for tissue in TARGET_TISSUES:
            params.append(("tissueSiteDetailId", tissue))

        page_cache = cache_path(cache_dir, "gtex_eqtl_page", f"{variant_id}__p{page}")
        payload, last_request_end, _ = fetch_json_with_retry(
            session,
            f"{GTEX_API_BASE}/association/singleTissueEqtl",
            params=params,
            cache_file=page_cache,
            min_interval_sec=GTEX_DELAY_SEC,
            max_retries=MAX_RETRIES,
            last_request_end=last_request_end,
            cache_only=cache_only,
        )
        if payload is None:
            break
        data = payload.get("data") or []
        if isinstance(data, list):
            hits.extend(data)
        paging = payload.get("paging_info") or {}
        number_of_pages = int(paging.get("numberOfPages", 1))
        page += 1

    # Defensive filter: API may ignore tissue filters on some pages.
    target_set = set(TARGET_TISSUES)
    hits = [hit for hit in hits if hit.get("tissueSiteDetailId") in target_set]
    aggregate_cache.parent.mkdir(parents=True, exist_ok=True)
    with aggregate_cache.open("w", encoding="utf-8") as handle:
        json.dump({"variantId": variant_id, "hits": hits}, handle, indent=2)
    return hits, last_request_end


def annotate_variants(
    variants: pl.DataFrame,
    cache_dir: Path,
    local_vep_path: Path,
    cache_only: bool,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Run GTEx + VEP annotation for every prioritized variant.

    Builds a per-variant summary (VEP + best eQTL) and a long eQTL table (one
    row per SNP–gene–tissue association). GTEx ``nes`` is exported as ``slope``.

    Args:
        variants: Prioritized variant DataFrame.
        cache_dir: API JSON cache directory.
        local_vep_path: Optional local VEP JSONL path.
        cache_only: Skip live API calls on cache miss.

    Returns:
        Tuple of ``(vep_gtex_summary_df, gtex_eqtl_long_df)``.
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_vep = load_local_vep(local_vep_path)

    session = requests.Session()
    session.headers.update(
        {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": (
                "rogen-aging-july-annotation/1.0 "
                "(ROGEN; academic research; GTEx Portal + Ensembl REST)"
            ),
        }
    )

    vep_rows: list[JsonDict] = []
    eqtl_rows: list[JsonDict] = []
    last_request_end: float | None = None

    records = variants.to_dicts()
    try:
        for row in tqdm(records, desc="Annotating variants", unit="var"):
            rsid = row.get("rsid")
            logger.debug("Annotating {}", rsid or row["variant_key"])

            vep_fields, last_request_end = fetch_vep_for_variant(
                session, row, cache_dir, local_vep, last_request_end, cache_only
            )
            vep_rows.append({"variant_key": row["variant_key"], **vep_fields})

            gtex_variant, last_request_end = resolve_gtex_variant(
                session, row, cache_dir, last_request_end, cache_only
            )
            gtex_variant_id = gtex_variant["variantId"] if gtex_variant else None
            n_hits = 0
            best_p: float | None = None
            best_tissue: str | None = None
            best_slope: float | None = None
            best_gene: str | None = None
            tissues: list[str] = []

            if gtex_variant_id:
                hits, last_request_end = fetch_gtex_eqtls(
                    session, gtex_variant_id, cache_dir, last_request_end, cache_only
                )
                n_hits = len(hits)
                for hit in hits:
                    tissue = str(hit.get("tissueSiteDetailId", ""))
                    slope = hit.get("nes")
                    p_value = hit.get("pValue")
                    gene_symbol = str(hit.get("geneSymbol", "") or "")
                    slope_f = float(slope) if slope is not None else float("nan")
                    p_f = float(p_value) if p_value is not None else float("nan")
                    eqtl_rows.append(
                        {
                            "chrom": row["chrom"],
                            "pos": row["pos"],
                            "ref": row["ref"],
                            "alt": row["alt"],
                            "rsid": rsid,
                            "input_gene_symbol": row.get("gene_symbol"),
                            "gtex_variant_id": gtex_variant_id,
                            "eqtl_gene_symbol": gene_symbol,
                            "tissue": tissue,
                            "slope": slope_f,
                            "p_value": p_f,
                        }
                    )
                    if tissue:
                        tissues.append(tissue)
                    # Track the most significant eQTL for the master summary row.
                    if p_value is not None and (best_p is None or float(p_value) < best_p):
                        best_p = float(p_value)
                        best_tissue = tissue
                        best_slope = slope_f
                        best_gene = gene_symbol

            vep_rows[-1].update(
                {
                    "gtex_variant_id": gtex_variant_id,
                    "gtex_n_eqtls": n_hits,
                    "gtex_best_tissue": best_tissue,
                    "gtex_best_gene": best_gene,
                    "gtex_best_slope": best_slope,
                    "gtex_best_p_value": best_p,
                    "gtex_tissues": ";".join(sorted(set(tissues))) if tissues else None,
                }
            )
    finally:
        session.close()

    vep_df = pl.DataFrame(vep_rows) if vep_rows else pl.DataFrame({"variant_key": []})
    eqtl_df = (
        pl.DataFrame(eqtl_rows)
        if eqtl_rows
        else pl.DataFrame(
            schema={
                "chrom": pl.Utf8,
                "pos": pl.Int64,
                "ref": pl.Utf8,
                "alt": pl.Utf8,
                "rsid": pl.Utf8,
                "input_gene_symbol": pl.Utf8,
                "gtex_variant_id": pl.Utf8,
                "eqtl_gene_symbol": pl.Utf8,
                "tissue": pl.Utf8,
                "slope": pl.Float64,
                "p_value": pl.Float64,
            }
        )
    )
    return vep_df, eqtl_df


def build_master_table(
    scores: pl.DataFrame,
    vep_gtex: pl.DataFrame,
) -> pl.DataFrame:
    """Combine variant, score, VEP, and GTEx summary columns.

    Args:
        scores: Variant table with AlphaGenome / AlphaMissense columns.
        vep_gtex: Per-variant VEP + GTEx summary from ``annotate_variants``.

    Returns:
        Column-ordered master annotation table (one row per variant).
    """
    master = scores.join(vep_gtex, on="variant_key", how="left")
    preferred = [
        "chrom",
        "pos",
        "ref",
        "alt",
        "rsid",
        "gene_symbol",
        "vep_most_severe_consequence",
        "vep_impact",
        "vep_transcript_id",
        "vep_gene_symbol",
        "vep_consequence_terms",
        "sift_prediction",
        "sift_score",
        "polyphen_prediction",
        "polyphen_score",
        "hgvsg",
        "hgvsc",
        "hgvsp",
        "alphagenome_ref_score",
        "alphagenome_alt_score",
        "alphagenome_diff",
        "alphagenome_perc_change",
        "alphagenome_abs_perc_change",
        "alphamissense_score",
        "alphamissense_class",
        "gtex_variant_id",
        "gtex_n_eqtls",
        "gtex_best_tissue",
        "gtex_best_gene",
        "gtex_best_slope",
        "gtex_best_p_value",
        "gtex_tissues",
        "variant_key",
    ]
    ordered = [c for c in preferred if c in master.columns]
    ordered.extend([c for c in master.columns if c not in ordered])
    return master.select(ordered)


def filter_high_impact(master: pl.DataFrame) -> pl.DataFrame:
    """Keep variants with AlphaMissense > 0.5 or VEP HIGH/MODERATE impact.

    Args:
        master: Combined master annotation table.

    Returns:
        Filtered high-impact subset (may be empty).
    """
    am_col = "alphamissense_score"
    impact_col = "vep_impact"
    exprs: list[pl.Expr] = []
    if am_col in master.columns:
        exprs.append(pl.col(am_col).cast(pl.Float64, strict=False) > ALPHAMISSENSE_HIGH_THRESHOLD)
    if impact_col in master.columns:
        exprs.append(pl.col(impact_col).cast(pl.Utf8).str.to_uppercase().is_in(list(HIGH_IMPACT_LEVELS)))
    if not exprs:
        logger.warning("No AlphaMissense or VEP impact columns available for high-impact filter")
        return master.head(0)
    mask = exprs[0]
    for expr in exprs[1:]:
        mask = mask | expr
    return master.filter(mask.fill_null(False))


def style_worksheet(ws: Worksheet, highlight: bool = False) -> None:
    """Apply header styling and freeze panes to an openpyxl worksheet.

    Args:
        ws: Target worksheet.
        highlight: When True, tint data rows (used for high-impact sheet).
    """
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(12, max_len + 2), 48)
    if highlight and ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = HIGHLIGHT_FILL


def write_excel_workbook(
    output_path: Path,
    master: pl.DataFrame,
    high_impact: pl.DataFrame,
    gtex_summary: pl.DataFrame,
) -> None:
    """Write the three-sheet supplementary Excel workbook via openpyxl.

    Sheets:
        1. ``Combined_Master`` — all variants with GTEx, VEP, and Alpha scores.
        2. ``High_Impact_Functional`` — AM > 0.5 or VEP HIGH/MODERATE.
        3. ``GTEx_eQTL_Summary`` — long tissue-specific eQTL table.

    Args:
        output_path: Destination ``.xlsx`` path.
        master: Combined master table.
        high_impact: Filtered high-impact subset.
        gtex_summary: Long-format GTEx eQTL table.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    sheets: list[tuple[str, pl.DataFrame, bool]] = [
        ("Combined_Master", master, False),
        ("High_Impact_Functional", high_impact, True),
        ("GTEx_eQTL_Summary", gtex_summary, False),
    ]

    for index, (title, frame, highlight) in enumerate(sheets):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title
        pdf = frame.to_pandas()
        for row in dataframe_to_rows(pdf, index=False, header=True):
            ws.append(row)
        style_worksheet(ws, highlight=highlight)

    wb.save(output_path)
    logger.info("Wrote workbook: {}", output_path)


@app.command()
def main(
    variants: Path = typer.Option(
        DEFAULT_VARIANTS,
        "--variants",
        help="Prioritized variant CSV (chrom,pos,ref,alt,rsid,gene_symbol).",
    ),
    alphagenome: Path = typer.Option(
        DEFAULT_ALPHAGENOME,
        "--alphagenome",
        help="Pre-computed AlphaGenome score parquet.",
    ),
    alphamissense: Path = typer.Option(
        DEFAULT_ALPHAMISSENSE,
        "--alphamissense",
        help="Pre-computed AlphaMissense score parquet.",
    ),
    output: Path = typer.Option(
        DEFAULT_OUTPUT,
        "--output",
        "-o",
        help="Output multi-sheet Excel path.",
    ),
    cache_dir: Path = typer.Option(
        DEFAULT_CACHE_DIR,
        "--cache-dir",
        help="JSON response cache directory for GTEx/VEP.",
    ),
    local_vep: Path = typer.Option(
        DEFAULT_LOCAL_VEP,
        "--local-vep",
        help="Optional local VEP JSONL (skips Ensembl for matching keys).",
    ),
    cache_only: bool = typer.Option(
        False,
        "--cache-only",
        help="Use disk cache only; do not call live APIs on cache miss.",
    ),
    demo: bool = typer.Option(
        False,
        "--demo",
        help=(
            "OFFLINE SMOKE TEST ONLY: write fixtures and run cache-only under "
            "outputs/demo/. Not used on the default production path."
        ),
    ),
    skip_manifest_check: bool = typer.Option(
        False,
        "--skip-manifest-check",
        help="Skip INPUT_MANIFEST.md required-file verification (ad-hoc runs).",
    ),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Debug logging."),
) -> None:
    """Execute the July functional annotation pipeline.

    Orchestrates score joins, GTEx/VEP annotation, high-impact filtering, and
    Excel export. The default path is **production**: all 47 prioritized variants
    → ``outputs/Supplementary_Table_1_Annotated_Variants.xlsx``. Network access
    is required on first run unless ``--cache-only`` is used with a warm cache.

    Args:
        variants: Prioritized variant CSV path.
        alphagenome: AlphaGenome score parquet path.
        alphamissense: AlphaMissense score parquet path.
        output: Output Excel workbook path.
        cache_dir: Directory for cached GTEx/VEP JSON responses.
        local_vep: Optional local VEP JSONL.
        cache_only: Do not call live APIs on cache miss.
        demo: Write fixtures and force an offline demo run under ``outputs/demo/``.
        skip_manifest_check: Bypass ``INPUT_MANIFEST.md`` checks.
        verbose: Enable DEBUG logging.
    """
    configure_logging(verbose)
    if demo:
        from rogen_aging.pipeline_fixtures import write_july_fixtures

        fixtures = write_july_fixtures(repo_root=REPO_ROOT)
        variants = fixtures["variants"]
        alphagenome = fixtures["alphagenome"]
        alphamissense = fixtures["alphamissense"]
        local_vep = fixtures["local_vep"]
        cache_dir = fixtures["cache_dir"]
        cache_only = True
        # Demo never overwrites the production Supplementary Table path.
        output = REPO_ROOT / "outputs" / "demo" / "Supplementary_Table_1_Annotated_Variants.xlsx"
        logger.info("Demo mode: fixtures written; running cache-only → {}", output)
    else:
        if not skip_manifest_check:
            verify_july_input_manifest()
        # Production always lands in outputs/ (never outputs/demo/).
        if "outputs/demo" in str(output).replace("\\", "/"):
            raise ValueError(
                f"Production output must not use outputs/demo/: {output}. "
                f"Expected {DEFAULT_OUTPUT}"
            )
        if output == DEFAULT_OUTPUT.parent / "demo" / DEFAULT_OUTPUT.name:
            output = DEFAULT_OUTPUT

    logger.info("Starting July annotation pipeline (production={} variants expected when using variants_47)", EXPECTED_VARIANT_COUNT)
    logger.info("GTEx dataset: {} | tissues: {}", GTEX_DATASET_ID, len(TARGET_TISSUES))
    logger.info("Input variants: {} | output: {}", variants, output)

    variant_df = load_prioritized_variants(variants)
    ag_df = load_score_matrix(alphagenome, "alphagenome")
    am_df = load_score_matrix(alphamissense, "alphamissense")
    scored = join_scores(variant_df, ag_df, am_df)

    # Prefer live Ensembl for production when local JSONL is a tiny demo fixture.
    if not demo and local_vep == DEFAULT_LOCAL_VEP and local_vep.is_file():
        local_lines = sum(1 for line in local_vep.open(encoding="utf-8") if line.strip())
        if local_lines < EXPECTED_VARIANT_COUNT:
            logger.warning(
                "Ignoring incomplete local VEP fixture {} ({} lines < {}); using Ensembl/cache",
                local_vep,
                local_lines,
                EXPECTED_VARIANT_COUNT,
            )
            local_vep = REPO_ROOT / "data" / "processed" / "_no_local_vep.jsonl"

    vep_gtex, gtex_long = annotate_variants(
        variant_df,
        cache_dir=cache_dir,
        local_vep_path=local_vep,
        cache_only=cache_only,
    )
    master = build_master_table(scored, vep_gtex)
    high_impact = filter_high_impact(master)

    # Halt on unexpected losses relative to the input set.
    if not demo and master.height != variant_df.height:
        raise RuntimeError(
            f"Master table has {master.height} rows but input had {variant_df.height}"
        )
    missing_vep = (
        master.filter(pl.col("vep_most_severe_consequence").is_null())["rsid"].to_list()
        if "vep_most_severe_consequence" in master.columns
        else []
    )
    if not demo and missing_vep:
        logger.error("VEP consequence missing for {} variants: {}", len(missing_vep), missing_vep)
        raise RuntimeError(
            f"Unexpected null VEP consequences for {len(missing_vep)} variants: {missing_vep}"
        )

    write_excel_workbook(output, master, high_impact, gtex_long)

    logger.info(
        "Done | variants={} | high_impact={} | gtex_eqtl_rows={} | output={}",
        master.height,
        high_impact.height,
        gtex_long.height,
        output,
    )


if __name__ == "__main__":
    app()
