#!/usr/bin/env python3
"""Publication-quality summary of predicted variant functional consequences.

Default MODE renders a one-page multi-panel figure (class donut, consequence
bars, gene × consequence matrix). Alternate modes: table list, standalone stacked bar.
Saves PNG (300 DPI), vector PDF, and an editable Excel workbook.
"""

from __future__ import annotations

import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Patch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Editable constants
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent.parent

INPUT_CSV = REPO_ROOT / "analysis" / "variant_consequence_predictions.csv"
OUTPUT_STEM = REPO_ROOT / "figures" / "variant_consequence_summary"

MODE = "summary"  # "summary" | "table" | "stacked_bar"

FIGURE_DPI = 300
PAGE_FIGSIZE = (11.0, 8.5)  # US letter, landscape — one page

REQUIRED_COLUMNS = ("gene", "rsid", "consequence", "class")
CLASS_ORDER = ("protein-altering", "non-coding")
ALLOWED_CLASSES = frozenset(CLASS_ORDER)

CLASS_COLORS = {
    "protein-altering": "#f5ddd6",
    "non-coding": "#dde8f0",
}
CLASS_EDGE_COLORS = {
    "protein-altering": "#c45c3e",
    "non-coding": "#4a7c9b",
}
SECTION_HEADER_COLOR = "#e8e8e8"
COLUMN_HEADER_COLOR = "#f0f0f0"
TABLE_CAPTION = (
    "Predicted functional consequences for longevity-associated variants "
    "across 41 genes (GRCh38)."
)
STACKED_BAR_CAPTION = (
    "Variant counts by functional consequence, split by protein-altering "
    "vs non-coding classes."
)

FONT_SIZES = {
    "base": 10,
    "panel_title": 11,
    "table_header": 10,
    "table_cell": 9,
    "section_header": 10,
    "caption": 9,
    "axis_label": 10,
    "tick": 8,
    "legend": 9,
    "heatmap_gene": 7.5,
    "heatmap_count": 7,
}

CONSEQUENCE_ORDER = (
    "missense_variant",
    "synonymous_variant",
    "intron_variant",
    "3_prime_utr_variant",
    "5_prime_utr_variant",
    "upstream_gene_variant",
    "downstream_gene_variant",
)

CONSEQUENCE_SHORT: dict[str, str] = {
    "missense_variant": "Missense",
    "synonymous_variant": "Synonymous",
    "intron_variant": "Intron",
    "3_prime_utr_variant": "3′ UTR",
    "5_prime_utr_variant": "5′ UTR",
    "upstream_gene_variant": "Upstream",
    "downstream_gene_variant": "Downstream",
    "intergenic_variant": "Intergenic",
    "regulatory_region_variant": "Regulatory",
}

CONSEQUENCE_DOT_COLORS: dict[str, str] = {
    "missense_variant": "#c45c3e",
    "synonymous_variant": "#e07a5f",
    "intron_variant": "#4a7c9b",
    "3_prime_utr_variant": "#6a9fb5",
    "5_prime_utr_variant": "#81b29a",
    "upstream_gene_variant": "#3d5a80",
    "downstream_gene_variant": "#5c677d",
}

TABLE_COLUMN_LABELS = ("Gene", "rsID", "Consequence", "Class")
TABLE_ROW_HEIGHT = 0.28
TABLE_MULTILINE_ROW_HEIGHT = 0.22
TABLE_FIGURE_WIDTH = 9.0
RSID_JOIN = ", "

EXCEL_COLUMN_WIDTHS = (14, 42, 36, 18)
EXCEL_ROW_HEIGHT = 15
EXCEL_ROW_HEIGHT_MULTILINE = 18


def _consequence_label(term: str) -> str:
    return CONSEQUENCE_SHORT.get(term, term.replace("_variant", "").replace("_", " ").title())


def _load_data(path: Path) -> pd.DataFrame:
    if not path.is_file():
        print(f"Error: input CSV not found: {path}", file=sys.stderr)
        sys.exit(1)

    df = pd.read_csv(path)
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        print(
            f"Error: input CSV is missing required column(s): {', '.join(missing)}",
            file=sys.stderr,
        )
        print(f"  Expected columns: {', '.join(REQUIRED_COLUMNS)}", file=sys.stderr)
        print(f"  Found columns: {', '.join(df.columns.astype(str))}", file=sys.stderr)
        sys.exit(1)

    work = df.loc[:, list(REQUIRED_COLUMNS)].copy()
    for col in REQUIRED_COLUMNS:
        work[col] = work[col].astype(str).str.strip()

    empty_mask = work.apply(lambda row: all(val in ("", "nan") for val in row), axis=1)
    work = work.loc[~empty_mask].reset_index(drop=True)
    if work.empty:
        print("Error: input CSV contains no variant rows.", file=sys.stderr)
        sys.exit(1)

    invalid = sorted(set(work["class"]) - ALLOWED_CLASSES)
    if invalid:
        print(
            f"Error: unexpected class value(s): {', '.join(invalid)}",
            file=sys.stderr,
        )
        print(
            f"  Allowed values: {', '.join(CLASS_ORDER)}",
            file=sys.stderr,
        )
        sys.exit(1)

    class_rank = {label: idx for idx, label in enumerate(CLASS_ORDER)}
    work["_class_rank"] = work["class"].map(class_rank)
    work = work.sort_values(["_class_rank", "gene", "rsid"], kind="stable")
    work = work.drop(columns=["_class_rank"]).reset_index(drop=True)
    return work


def _apply_rcparams() -> None:
    plt.rcParams.update(
        {
            "font.size": FONT_SIZES["base"],
            "font.family": "sans-serif",
            "axes.facecolor": "white",
            "figure.facecolor": "white",
        }
    )


def _save_figure(fig: plt.Figure, output_stem: Path) -> None:
    output_stem = Path(output_stem)
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    png_path = output_stem.with_suffix(".png")
    pdf_path = output_stem.with_suffix(".pdf")
    fig.savefig(png_path, dpi=FIGURE_DPI, bbox_inches="tight")
    fig.savefig(pdf_path, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved: {png_path}")
    print(f"Saved: {pdf_path}")


def _class_counts(df: pd.DataFrame) -> dict[str, int]:
    counts = df["class"].value_counts()
    return {label: int(counts.get(label, 0)) for label in CLASS_ORDER}


def _format_section_title(class_label: str, variant_count: int, gene_count: int) -> str:
    display = class_label.replace("-", " ").title()
    return f"{display} — {gene_count} genes, {variant_count} variants"


def _format_consequence(rsids: list[str], consequences: list[str]) -> str:
    if len(rsids) == 1:
        return consequences[0]
    unique = sorted(set(consequences))
    if len(unique) == 1:
        return f"{unique[0]} (×{len(rsids)})"
    return "\n".join(f"{rsid}: {cons}" for rsid, cons in zip(rsids, consequences, strict=True))


def _aggregate_by_gene(section_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for gene, grp in section_df.groupby("gene", sort=True):
        grp = grp.sort_values("rsid", kind="stable")
        rsids = grp["rsid"].tolist()
        consequences = grp["consequence"].tolist()
        gene_label = f"{gene} ({len(rsids)})" if len(rsids) > 1 else str(gene)
        rows.append(
            {
                "gene": gene_label,
                "rsid": RSID_JOIN.join(rsids),
                "consequence": _format_consequence(rsids, consequences),
                "class": grp["class"].iloc[0],
            }
        )
    return pd.DataFrame(rows)


def _heatmap_row_order(df: pd.DataFrame) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for class_label in CLASS_ORDER:
        genes = sorted(df.loc[df["class"] == class_label, "gene"].unique())
        rows.extend((gene, class_label) for gene in genes)
    return rows


def _build_gene_consequence_matrix(
    df: pd.DataFrame,
) -> tuple[np.ndarray, list[tuple[str, str]], list[str]]:
    row_order = _heatmap_row_order(df)
    col_labels = [_consequence_label(c) for c in CONSEQUENCE_ORDER]
    matrix = np.zeros((len(row_order), len(CONSEQUENCE_ORDER)), dtype=int)
    for row_idx, (gene, class_label) in enumerate(row_order):
        subset = df.loc[(df["gene"] == gene) & (df["class"] == class_label)]
        for col_idx, consequence in enumerate(CONSEQUENCE_ORDER):
            matrix[row_idx, col_idx] = int((subset["consequence"] == consequence).sum())
    return matrix, row_order, col_labels


def _plot_class_donut(ax: plt.Axes, df: pd.DataFrame) -> None:
    counts = _class_counts(df)
    values = [counts["protein-altering"], counts["non-coding"]]
    labels = [
        f"Protein-altering\n{values[0]} variants",
        f"Non-coding\n{values[1]} variants",
    ]
    colors = [CLASS_EDGE_COLORS["protein-altering"], CLASS_EDGE_COLORS["non-coding"]]
    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        colors=colors,
        startangle=90,
        counterclock=False,
        autopct=lambda pct: f"{pct:.0f}%",
        pctdistance=0.75,
        labeldistance=1.12,
        wedgeprops={"width": 0.48, "edgecolor": "white", "linewidth": 1.5},
        textprops={"fontsize": FONT_SIZES["tick"]},
    )
    for autotext in autotexts:
        autotext.set_color("white")
        autotext.set_fontsize(FONT_SIZES["tick"])
        autotext.set_fontweight("bold")
    n_genes = df["gene"].nunique()
    ax.text(
        0,
        0,
        f"{n_genes} genes\n{len(df)} variants",
        ha="center",
        va="center",
        fontsize=FONT_SIZES["panel_title"],
        fontweight="bold",
        color="#333333",
    )
    ax.set_title("Impact class", fontsize=FONT_SIZES["panel_title"], fontweight="bold", pad=8)


def _plot_consequence_bars(ax: plt.Axes, df: pd.DataFrame) -> None:
    counts = (
        df.groupby(["consequence", "class"], observed=True)
        .size()
        .reset_index(name="count")
    )
    consequences = [c for c in CONSEQUENCE_ORDER if c in set(counts["consequence"])]
    extra = sorted(set(counts["consequence"]) - set(consequences))
    consequences.extend(extra)

    y = np.arange(len(consequences))
    protein = []
    noncoding = []
    for consequence in consequences:
        subset = counts.loc[counts["consequence"] == consequence].set_index("class")
        protein.append(int(subset["count"].get("protein-altering", 0)))
        noncoding.append(int(subset["count"].get("non-coding", 0)))

    ax.barh(
        y,
        protein,
        height=0.72,
        color=CLASS_EDGE_COLORS["protein-altering"],
        label="Protein-altering",
        edgecolor="white",
        linewidth=0.6,
    )
    ax.barh(
        y,
        noncoding,
        left=protein,
        height=0.72,
        color=CLASS_EDGE_COLORS["non-coding"],
        label="Non-coding",
        edgecolor="white",
        linewidth=0.6,
    )
    ax.set_yticks(y)
    ax.set_yticklabels([_consequence_label(c) for c in consequences], fontsize=FONT_SIZES["tick"])
    ax.set_xlabel("Variant count", fontsize=FONT_SIZES["axis_label"])
    ax.set_title("Consequence type", fontsize=FONT_SIZES["panel_title"], fontweight="bold", pad=8)
    ax.invert_yaxis()
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.xaxis.grid(True, color="#dddddd", linewidth=0.8, alpha=0.8)
    ax.set_axisbelow(True)
    ax.legend(fontsize=FONT_SIZES["legend"], frameon=False, loc="lower right")
    ax.tick_params(axis="x", labelsize=FONT_SIZES["tick"])


def _plot_gene_consequence_matrix(ax: plt.Axes, df: pd.DataFrame) -> None:
    matrix, row_order, col_labels = _build_gene_consequence_matrix(df)
    n_rows, n_cols = matrix.shape

    ax.set_xlim(-0.5, n_cols - 0.5)
    ax.set_ylim(n_rows - 0.5, -0.5)
    ax.set_xticks(range(n_cols))
    ax.set_xticklabels(col_labels, fontsize=FONT_SIZES["tick"])
    ax.set_yticks(range(n_rows))
    ax.set_yticklabels([gene for gene, _ in row_order], fontsize=FONT_SIZES["heatmap_gene"])

    for tick, (_, class_label) in zip(ax.get_yticklabels(), row_order, strict=True):
        tick.set_color(CLASS_EDGE_COLORS[class_label])
        tick.set_fontweight("bold")

    ax.set_xlabel("Predicted consequence", fontsize=FONT_SIZES["axis_label"])
    ax.set_ylabel("Gene", fontsize=FONT_SIZES["axis_label"])
    ax.set_title(
        "Gene × consequence matrix (dot size = variant count)",
        fontsize=FONT_SIZES["panel_title"],
        fontweight="bold",
        pad=10,
    )

    for row_idx in range(n_rows + 1):
        ax.axhline(row_idx - 0.5, color="#e0e0e0", linewidth=0.6, zorder=0)
    for col_idx in range(n_cols + 1):
        ax.axvline(col_idx - 0.5, color="#e0e0e0", linewidth=0.6, zorder=0)

    protein_rows = sum(1 for _, cls in row_order if cls == "protein-altering")
    if 0 < protein_rows < n_rows:
        ax.axhline(protein_rows - 0.5, color="#888888", linewidth=1.4, zorder=1)

    max_count = int(matrix.max()) if matrix.size else 1
    size_scale = 220 if max_count <= 3 else 180

    for row_idx, (_, class_label) in enumerate(row_order):
        for col_idx, consequence in enumerate(CONSEQUENCE_ORDER):
            count = int(matrix[row_idx, col_idx])
            if count == 0:
                continue
            color = CONSEQUENCE_DOT_COLORS.get(consequence, CLASS_EDGE_COLORS[class_label])
            ax.scatter(
                col_idx,
                row_idx,
                s=size_scale * count,
                c=color,
                edgecolors="white",
                linewidths=0.8,
                zorder=2,
            )
            if count > 1:
                ax.text(
                    col_idx,
                    row_idx,
                    str(count),
                    ha="center",
                    va="center",
                    fontsize=FONT_SIZES["heatmap_count"],
                    color="white",
                    fontweight="bold",
                    zorder=3,
                )

    legend_handles = [
        Patch(facecolor=CLASS_EDGE_COLORS["protein-altering"], label="Protein-altering gene"),
        Patch(facecolor=CLASS_EDGE_COLORS["non-coding"], label="Non-coding gene"),
    ]
    ax.legend(
        handles=legend_handles,
        fontsize=FONT_SIZES["legend"],
        frameon=False,
        loc="upper left",
        bbox_to_anchor=(1.01, 1.0),
    )


def plot_summary(df: pd.DataFrame, output_stem: Path) -> None:
    _apply_rcparams()
    fig = plt.figure(figsize=PAGE_FIGSIZE, facecolor="white")
    gs = fig.add_gridspec(
        nrows=2,
        ncols=2,
        height_ratios=[1.0, 2.2],
        width_ratios=[1.0, 1.35],
        hspace=0.38,
        wspace=0.28,
        left=0.07,
        right=0.92,
        top=0.93,
        bottom=0.08,
    )

    _plot_class_donut(fig.add_subplot(gs[0, 0]), df)
    _plot_consequence_bars(fig.add_subplot(gs[0, 1]), df)
    _plot_gene_consequence_matrix(fig.add_subplot(gs[1, :]), df)

    fig.text(
        0.5,
        0.015,
        TABLE_CAPTION,
        ha="center",
        va="bottom",
        fontsize=FONT_SIZES["caption"],
        color="#666666",
        style="italic",
    )
    _save_figure(fig, output_stem)
    export_table_excel(df, output_stem)


def _row_line_count(row: list[str]) -> int:
    return max(text.count("\n") + 1 for text in row)


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _build_table_sections(df: pd.DataFrame) -> tuple[list[list[str]], list[str], list[str | None]]:
    variant_counts = _class_counts(df)
    table_rows: list[list[str]] = [list(TABLE_COLUMN_LABELS)]
    row_kinds: list[str] = ["column_header"]
    row_classes: list[str | None] = [None]

    for class_label in CLASS_ORDER:
        section_df = df.loc[df["class"] == class_label]
        if section_df.empty:
            continue
        grouped = _aggregate_by_gene(section_df)
        table_rows.append(
            [
                _format_section_title(
                    class_label,
                    variant_counts[class_label],
                    len(grouped),
                ),
                "",
                "",
                "",
            ]
        )
        row_kinds.append("section_header")
        row_classes.append(class_label)
        for gene, rsid, consequence, row_class in zip(
            grouped["gene"],
            grouped["rsid"],
            grouped["consequence"],
            grouped["class"],
            strict=True,
        ):
            table_rows.append([gene, rsid, consequence, row_class])
            row_kinds.append("data")
            row_classes.append(class_label)

    return table_rows, row_kinds, row_classes


def export_table_excel(df: pd.DataFrame, output_stem: Path) -> None:
    table_rows, row_kinds, row_classes = _build_table_sections(df)
    output_stem = Path(output_stem)
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_stem.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "summary_table"

    for col_idx, width in enumerate(EXCEL_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, (cells, kind, class_label) in enumerate(
        zip(table_rows, row_kinds, row_classes, strict=True),
        start=1,
    ):
        line_count = _row_line_count(cells)
        ws.row_dimensions[row_idx].height = (
            EXCEL_ROW_HEIGHT
            if line_count == 1
            else EXCEL_ROW_HEIGHT_MULTILINE * line_count
        )

        for col_idx, value in enumerate(cells, start=1):
            if kind == "section_header" and col_idx > 1:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if kind == "column_header" else "left",
                wrap_text=col_idx in (2, 3) or kind == "section_header",
            )

            if kind == "column_header":
                cell.font = Font(bold=True)
                cell.fill = _hex_fill(COLUMN_HEADER_COLOR)
                continue

            if kind == "section_header":
                cell.font = Font(bold=True)
                cell.fill = _hex_fill(SECTION_HEADER_COLOR)
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=4,
                )
                continue

            cell.fill = _hex_fill(CLASS_COLORS[class_label])
            if col_idx == 1:
                cell.font = Font(bold=True)

    caption_row = len(table_rows) + 2
    ws.cell(row=caption_row, column=1, value=TABLE_CAPTION)
    ws.merge_cells(
        start_row=caption_row,
        start_column=1,
        end_row=caption_row,
        end_column=4,
    )
    caption_cell = ws.cell(row=caption_row, column=1)
    caption_cell.font = Font(italic=True, size=9, color="444444")
    caption_cell.alignment = Alignment(wrap_text=True)

    variants = wb.create_sheet("variants")
    variants.append(list(REQUIRED_COLUMNS))
    for col_idx in range(1, len(REQUIRED_COLUMNS) + 1):
        cell = variants.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = _hex_fill(COLUMN_HEADER_COLOR)
    for _, row in df.iterrows():
        variants.append([row["gene"], row["rsid"], row["consequence"], row["class"]])
    variants.freeze_panes = "A2"
    for col_idx, width in enumerate((14, 16, 28, 18), start=1):
        variants.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")


def plot_table(df: pd.DataFrame, output_stem: Path) -> None:
    _apply_rcparams()
    table_rows, row_kinds, row_classes = _build_table_sections(df)
    row_heights: list[float] = [TABLE_ROW_HEIGHT]

    for cells, kind in zip(table_rows[1:], row_kinds[1:], strict=True):
        if kind == "section_header":
            row_heights.append(TABLE_ROW_HEIGHT)
        else:
            lines = _row_line_count(cells)
            row_heights.append(
                TABLE_ROW_HEIGHT if lines == 1 else TABLE_MULTILINE_ROW_HEIGHT * lines
            )

    fig_height = max(4.0, 0.9 + sum(row_heights))
    fig, ax = plt.subplots(figsize=(TABLE_FIGURE_WIDTH, fig_height))
    ax.axis("off")

    table = ax.table(
        cellText=table_rows,
        loc="upper center",
        cellLoc="left",
        colWidths=[0.12, 0.30, 0.40, 0.18],
    )
    table.auto_set_font_size(False)
    table.scale(1.0, 1.35)

    for (row_idx, col_idx), cell in table.get_celld().items():
        cell.set_edgecolor("#cccccc")
        cell.set_linewidth(0.5)
        kind = row_kinds[row_idx]
        class_label = row_classes[row_idx]

        if kind == "column_header":
            cell.set_facecolor(COLUMN_HEADER_COLOR)
            cell.set_text_props(weight="bold", fontsize=FONT_SIZES["table_header"])
            cell.get_text().set_ha("center")
            continue

        if kind == "section_header":
            cell.set_facecolor(SECTION_HEADER_COLOR)
            cell.set_text_props(weight="bold", fontsize=FONT_SIZES["section_header"])
            if col_idx > 0:
                cell.set_visible(False)
            else:
                cell.get_text().set_ha("left")
            continue

        cell.set_facecolor(CLASS_COLORS[class_label])
        cell.set_text_props(fontsize=FONT_SIZES["table_cell"])
        if col_idx == 0:
            cell.get_text().set_weight("bold")
        if col_idx == 3:
            cell.get_text().set_ha("center")
        if col_idx in (1, 2):
            cell.get_text().set_wrap(True)

    fig.text(
        0.5,
        0.02,
        TABLE_CAPTION,
        ha="center",
        va="bottom",
        fontsize=FONT_SIZES["caption"],
        color="#444444",
        wrap=True,
    )
    fig.subplots_adjust(left=0.04, right=0.98, top=0.98, bottom=0.06)
    _save_figure(fig, output_stem)
    export_table_excel(df, output_stem)


def _consequence_order(counts: pd.DataFrame) -> list[str]:
    totals = counts.groupby("consequence")["count"].sum().sort_values(ascending=False)
    return totals.index.tolist()


def plot_stacked_bar(df: pd.DataFrame, output_stem: Path) -> None:
    _apply_rcparams()

    counts = (
        df.groupby(["consequence", "class"], observed=True)
        .size()
        .reset_index(name="count")
    )
    consequences = _consequence_order(counts)
    x = np.arange(len(consequences))
    bar_width = 0.72

    protein_counts = []
    noncoding_counts = []
    for consequence in consequences:
        subset = counts.loc[counts["consequence"] == consequence].set_index("class")
        protein_counts.append(int(subset["count"].get("protein-altering", 0)))
        noncoding_counts.append(int(subset["count"].get("non-coding", 0)))

    fig, ax = plt.subplots(figsize=(max(8.0, len(consequences) * 0.55), 5.5), layout="constrained")

    ax.bar(
        x,
        protein_counts,
        bar_width,
        label="Protein-altering",
        color=CLASS_EDGE_COLORS["protein-altering"],
        edgecolor="white",
        linewidth=0.6,
    )
    ax.bar(
        x,
        noncoding_counts,
        bar_width,
        bottom=protein_counts,
        label="Non-coding",
        color=CLASS_EDGE_COLORS["non-coding"],
        edgecolor="white",
        linewidth=0.6,
    )

    ax.set_xticks(x)
    ax.set_xticklabels([_consequence_label(c) for c in consequences], rotation=45, ha="right", fontsize=FONT_SIZES["tick"])
    ax.set_ylabel("Variant count", fontsize=FONT_SIZES["axis_label"])
    ax.tick_params(axis="y", labelsize=FONT_SIZES["tick"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color="#dddddd", linewidth=0.8, linestyle="-", alpha=0.8)
    ax.legend(fontsize=FONT_SIZES["legend"], frameon=False, loc="upper right")
    fig.text(
        0.5,
        -0.02,
        STACKED_BAR_CAPTION,
        ha="center",
        va="top",
        fontsize=FONT_SIZES["caption"],
        color="#444444",
        transform=ax.transAxes,
    )
    _save_figure(fig, output_stem)


def main() -> None:
    df = _load_data(Path(INPUT_CSV))
    output_stem = Path(OUTPUT_STEM)

    if MODE == "summary":
        plot_summary(df, output_stem)
    elif MODE == "table":
        plot_table(df, output_stem)
    elif MODE == "stacked_bar":
        plot_stacked_bar(df, output_stem)
    else:
        print(
            f"Error: unknown MODE {MODE!r}; use 'summary', 'table', or 'stacked_bar'.",
            file=sys.stderr,
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
